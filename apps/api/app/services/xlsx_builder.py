"""WBS XLSX 빌더 (스켈레톤)."""
from typing import IO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_wbs_xlsx(out: IO[bytes], project_uuid: str, phases: int = 5) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"
    ws.cell(row=1, column=1, value=f"WBS — {project_uuid}").font = Font(bold=True, size=14)

    headers = ["Phase", "Task", "담당", "기간(일)", "산출물"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3A5F")
        cell.alignment = Alignment(horizontal="center")

    sample = [
        (1, "분석", "요건 정의", "PM", 5, "요구사항 명세서"),
        (2, "분석", "공고문 분석", "PM", 3, "분석 보고서"),
        (3, "설계", "아키텍처 설계", "TA", 7, "아키텍처 문서"),
        (4, "설계", "DB 설계", "DA", 5, "DB 설계서"),
        (5, "개발", "BE 개발", "BE", 20, "API"),
        (6, "개발", "FE 개발", "FE", 20, "Web UI"),
        (7, "테스트", "통합 테스트", "QA", 10, "테스트 결과서"),
        (8, "이행", "배포", "DevOps", 3, "배포 가이드"),
    ]
    r = 4
    for _, phase, task, owner, dur, deliv in sample:
        ws.cell(row=r, column=1, value=phase)
        ws.cell(row=r, column=2, value=task)
        ws.cell(row=r, column=3, value=owner)
        ws.cell(row=r, column=4, value=dur)
        ws.cell(row=r, column=5, value=deliv)
        r += 1

    for col, w in enumerate([10, 24, 12, 12, 36], start=1):
        ws.column_dimensions[chr(64 + col)].width = w

    wb.save(out)
